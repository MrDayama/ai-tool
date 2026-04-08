import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_spec_excel():
    output_path = "AI_Agent_Detailed_Specification.xlsx"
    
    # --- 1. システム基本設計 (Basic Design) ---
    vision_data = {
        "設計項目": ["システム名称", "基本理念", "対応言語", "モデル設定", "自律性"],
        "定義内容": [
            "AI Autonomous Agent (自律型AIエージェント)",
            "ユーザーの曖昧な目標を技術的タスクに昇華し、完遂するまで自己修正し続ける。",
            "Python 3.12+ / 日本語環境完全対応",
            "OpenAI gpt-4o (デフォルト設定)",
            "人間をループに介在させない自律的なTool UseとCriticレビュー"
        ]
    }

    # --- 2. モジュール・関数詳細 (Function Reference) ---
    function_data = {
        "モジュール名": [
            "ManagerAgent", "ManagerAgent", "PlannerAgent", "PlannerAgent",
            "ExecutorAgent", "ExecutorAgent", "CriticAgent", "MemoryStore",
            "GitHubTool", "CodeRunner"
        ],
        "関数名": [
            "start_goal()", "run_cli()", "plan()", "refine_plan()", "execute()", 
            "self.tools", "review()", "get_full_context()", "create_issue()", "run_python_code()"
        ],
        "機能説明": [
            "目標の受け取り、プランナー・実行・評価のループ制御とログ出力。",
            "対話型のコマンドラインインターフェースを起動。",
            "目標をLLMに渡し、JSON形式のタスクリストを生成させる。",
            "批評家(Critic)のフィードバックを受け、残タスクを再計画(Improve)。",
            "プランから抽出されたタスクに基づき、適切なツールを選択して動作させる。",
            "Web/Python/File/Shell/GitHub ツールを辞書管理し呼び出す。",
            "実行結果が要件を満たしているか、パス・不合格を厳密に判定。",
            "現在のタスク状況、実行履歴、残タスクを合算し、エージェントへ文脈として渡す。",
            "GitHub APIを使用して指定リポジトリにIssue(課題)を自動起票。",
            "一時ファイルを作成し、サブプロセス経由でPythonコードを安全に実行・結果取得。"
        ]
    }

    # --- 3. 使い方（セットアップ・運用） ---
    usage_data = {
        "工程": ["1. 準備", "2. 構築", "3. 設定", "4. 運用", "5. 配布"],
        "詳細ステップ": [
            "OpenAI APIキーを用意。GitHub操作が必要ならトークンも発行。",
            "python setup.py を実行。venvが作成されライブラリが自動導入される。",
            ".envファイルを作成。OPENAI_API_KEY=sk-... と記述。",
            "python main.py を起動。日本語で『GitHubリポジトリを分析して』等と入力。",
            "python build_exe.py を実行。dist/内に1ファイルに固まったexeが出現。"
        ],
        "トラブルシューティング": [
            "OpenAIの残高確認", "pythonのパス確認", "APIキーの権限確認", "日本語入力バッファの確認", "exeと同じ場所に設定ファイルがあるか"
        ]
    }

    # --- Excel Construction ---
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        pd.DataFrame(vision_data).to_excel(writer, sheet_name='1.基本設計', index=False)
        pd.DataFrame(function_data).to_excel(writer, sheet_name='2.機能・関数定義', index=False)
        pd.DataFrame(usage_data).to_excel(writer, sheet_name='3.マニュアル', index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            # Column Sizing & Style
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 50
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = thin_border
            # Header Style
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

        # --- Embedding Japanese Diagrams ---
        if os.path.exists("arch_jp.png"):
            img = Image("arch_jp.png")
            img.width, img.height = (480, 480)
            workbook['1.基本設計'].add_image(img, 'D2')

        if os.path.exists("github_jp.png"):
            img = Image("github_jp.png")
            img.width, img.height = (430, 430)
            workbook['2.機能・関数定義'].add_image(img, 'D5')

    print(f"High-detailed Japanese Excel spec created: {output_path}")

if __name__ == "__main__":
    create_spec_excel()
