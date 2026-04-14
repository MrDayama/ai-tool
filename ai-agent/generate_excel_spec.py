import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_spec_excel():
    output_path = "AI_Agent_Specification_v2.xlsx"
    
    # --- Data Preparation ---
    # 1. Overview & Vision
    vision_data = {
        "セクション": ["コンセプト", "主要ターゲット", "開発スタック", "GitHub連携", "配布形態"],
        "内容": [
            "完全自律型のAIエンジニア・アシスタント。目標からタスクを自己分解し、評価・修正を繰り返します。",
            "エンジニア、データサイエンティスト、自動化を求めるビジネスユーザー",
            "Python 3.12 / OpenAI API (GPT-4) / Rich CLI",
            "APIを通じたリポジトリ操作、Issue管理、コード読み取りに対応",
            "Pythonソース。およびPyInstallerによる単体実行ファイル(.exe)"
        ]
    }
    
    # 2. Detailed Architecture Design
    design_data = {
        "コンポーネント": ["Manager (司令塔)", "Planner (計画者)", "Executor (実行者)", "Critic (評価者)", "Memory (記憶)"],
        "詳細設計内容": [
            "全体のライフサイクルを管理。進捗を監視し、フェーズ間の遷移を制御します。",
            "LLMを用いて再帰的にタスクを分解。複雑な依存関係をリスト化します。",
            "Tool Use機能を備え、外部（Web, GitHub, OS）と直接対話します。",
            "結果の品質を厳格にチェック。失敗時はPlannerに差し戻し、自己修復を試みます。",
            "JSONベースの永続化。過去の成功・失敗事例をコンテキストとして保持。"
        ],
        "動作ループ": ["GOAL入力", "PLAN生成", "EXECUTE (ツール使用)", "CRITIC (評価)", "IMPROVE / NEXT"]
    }

    # 3. Step-by-Step Usage
    usage_data = {
        "ステップ": ["1. 準備", "2. セットアップ", "3. 連携設定", "4. AI対話開始", "5. 配布用ビルド"],
        "作業内容": [
            "OpenAI APIキーとGitHubトークンを取得",
            "setup.pyを実行し仮想環境を構築",
            ".envファイルに必要なAPIキーを書き込む",
            "main.pyを起動し、日本語で目標を入力（例：天気予報を要約して）",
            "build_exe.pyを実行。distフォルダにexeが生成される"
        ],
        "注意点": ["必須", "一回のみ", "GITHUB_TOKENは任意", "日本語対応", "Windows/Mac対応"]
    }

    # --- Excel Construction ---
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        pd.DataFrame(vision_data).to_excel(writer, sheet_name='1. システム概要', index=False)
        pd.DataFrame(design_data).to_excel(writer, sheet_name='2. 詳細設計', index=False)
        pd.DataFrame(usage_data).to_excel(writer, sheet_name='3. 使い方ガイド', index=False)

        # Style & Images
        workbook = writer.book
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            # Format Columns
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 40
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = thin_border
            # Format Header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

        # --- Embed Images ---
        # Architecture Diagram to Design Sheet
        if os.path.exists("architecture_diagram.png"):
            img_design = Image("architecture_diagram.png")
            img_design.width, img_design.height = (400, 400) # Resize
            workbook['2. 詳細設計'].add_image(img_design, 'D2')

        # GitHub Diagram to Overview Sheet
        if os.path.exists("github_diagram.png"):
            img_github = Image("github_diagram.png")
            img_github.width, img_github.height = (400, 400)
            workbook['1. システム概要'].add_image(img_github, 'D2')

    print(f"Detailed Excel spec with images created: {output_path}")

if __name__ == "__main__":
    create_spec_excel()
